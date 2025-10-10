import openpyxl
import json

wb = openpyxl.load_workbook('creative_traits_framework_full.xlsx')
ws = wb.active

# Get headers from row 1
headers = []
for cell in ws[1]:
    if cell.value:
        headers.append(cell.value)

# Extract options for each column
traits = {}
for col_idx, header in enumerate(headers):
    options = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        value = row[col_idx] if col_idx < len(row) else None
        if value and str(value).strip() and value not in options:
            options.append(str(value).strip())
    traits[header] = options

print(json.dumps(traits, indent=2))
